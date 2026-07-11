from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "已制作.xlsx"
OUT_DIR = ROOT / "data" / "cards_raw"
REPORT = ROOT / "docs" / "excel-import-report.md"


SHEET_KEYS = {
    "战斗人物": "combat_characters",
    "附加人物": "attached_characters",
    "物品": "items",
    "称号": "titles",
    "场景": "scenes",
    "废弃": "deprecated",
}


FIELD_KEYS = {
    "名称": "title",
    "血量": "life",
    "身份": "identity",
    "描述": "description",
    "关系": "relationships",
    "兵器": "weapons",
    "出处": "source_work",
    "作者": "author_group",
    "性别": "gender",
    "特性": "traits",
    "类别": "item_category",
}


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.replace("\r\n", "\n").replace("\r", "\n")
        return normalized.strip()
    return value


def is_blank(value: Any) -> bool:
    value = clean_cell(value)
    return value is None or value == ""


def stable_id(sheet: str, row_number: int, title: str | None) -> str:
    seed = f"{sheet}|{row_number}|{title or ''}"
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:12]
    return f"raw_{digest}"


def normalize_title(title: str | None) -> str:
    if not title:
        return ""
    return re.sub(r"\s+", "", title.replace("（", "(").replace("）", ")"))


def read_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    header_values = [clean_cell(cell.value) for cell in ws[1]]
    field_names: list[str] = []
    unnamed_count = 0
    for idx, header in enumerate(header_values, start=1):
        if isinstance(header, str) and header:
            field_names.append(header)
        else:
            unnamed_count += 1
            field_names.append(f"__unnamed_{idx}")

    records: list[dict[str, Any]] = []
    rows_with_extra_values: list[int] = []
    blank_name_rows: list[int] = []

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [clean_cell(value) for value in row]
        if all(is_blank(value) for value in values):
            continue

        raw_fields = {
            field_names[i]: values[i] if i < len(values) else None
            for i in range(len(field_names))
        }
        raw_fields = {key: value for key, value in raw_fields.items() if not is_blank(value)}

        title = raw_fields.get("名称")
        title = title if isinstance(title, str) else str(title) if title is not None else None
        if not title:
            blank_name_rows.append(row_number)

        unnamed_values = {
            key: value
            for key, value in raw_fields.items()
            if key.startswith("__unnamed_") and not is_blank(value)
        }
        if unnamed_values:
            rows_with_extra_values.append(row_number)

        fields: dict[str, Any] = {}
        extra_fields: dict[str, Any] = {}
        for key, value in raw_fields.items():
            mapped = FIELD_KEYS.get(key)
            if mapped:
                fields[mapped] = value
            else:
                extra_fields[key] = value

        title_value = fields.get("title")
        record = {
            "id": stable_id(ws.title, row_number, title_value if isinstance(title_value, str) else None),
            "source": {
                "workbook": SOURCE.name,
                "sheet": ws.title,
                "row": row_number,
            },
            "category": SHEET_KEYS.get(ws.title, ws.title),
            "title": title_value,
            "normalized_title": normalize_title(title_value if isinstance(title_value, str) else None),
            "fields": fields,
            "extra_fields": extra_fields,
            "raw_fields": raw_fields,
        }
        records.append(record)

    stats = {
        "sheet": ws.title,
        "category": SHEET_KEYS.get(ws.title, ws.title),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header_values": header_values,
        "record_count": len(records),
        "unnamed_header_count": unnamed_count,
        "rows_with_extra_values": rows_with_extra_values,
        "blank_name_rows": blank_name_rows,
    }
    return records, stats


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as fh:
        for record in records:
            fh.write(json.dumps(record, ensure_ascii=False, default=str))
            fh.write("\n")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    all_records: list[dict[str, Any]] = []
    sheet_stats: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        records, stats = read_sheet(ws)
        all_records.extend(records)
        sheet_stats.append(stats)
        key = SHEET_KEYS.get(ws.title, ws.title)
        write_jsonl(OUT_DIR / f"{key}.jsonl", records)

    write_jsonl(OUT_DIR / "all_cards.jsonl", all_records)
    (OUT_DIR / "manifest.json").write_text(
        json.dumps(
            {
                "source_workbook": SOURCE.name,
                "imported_at_utc": datetime.now(timezone.utc).isoformat(),
                "record_count": len(all_records),
                "sheets": sheet_stats,
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )

    title_counter = Counter(
        record["title"] for record in all_records if isinstance(record.get("title"), str)
    )
    duplicates = {title: count for title, count in title_counter.items() if count > 1}
    by_author = Counter(
        record["fields"].get("author_group") for record in all_records if record["fields"].get("author_group")
    )
    by_category = Counter(record["category"] for record in all_records)

    duplicate_locations: dict[str, list[str]] = defaultdict(list)
    for record in all_records:
        title = record.get("title")
        if title in duplicates:
            source = record["source"]
            duplicate_locations[title].append(f"{source['sheet']}!{source['row']}")

    lines = [
        "# Excel 导入报告",
        "",
        f"- 来源：`{SOURCE.name}`",
        f"- 输出目录：`data/cards_raw/`",
        f"- 总记录数：{len(all_records)}",
        "",
        "## Sheet 统计",
        "",
        "| Sheet | 分类 | 记录数 | 最大行 | 最大列 | 未命名表头数 | 有未命名列内容的行数 | 空名称行数 |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for stats in sheet_stats:
        lines.append(
            f"| {stats['sheet']} | `{stats['category']}` | {stats['record_count']} | "
            f"{stats['max_row']} | {stats['max_column']} | {stats['unnamed_header_count']} | "
            f"{len(stats['rows_with_extra_values'])} | {len(stats['blank_name_rows'])} |"
        )

    lines.extend(["", "## 分类统计", ""])
    for category, count in sorted(by_category.items()):
        lines.append(f"- `{category}`: {count}")

    lines.extend(["", "## 作者统计", ""])
    for author, count in by_author.most_common():
        lines.append(f"- {author}: {count}")

    lines.extend(["", "## 重名标题", ""])
    if duplicate_locations:
        for title, locations in sorted(duplicate_locations.items()):
            lines.append(f"- {title}: {', '.join(locations)}")
    else:
        lines.append("- 无")

    lines.extend(["", "## 需要人工查看的结构问题", ""])
    any_issue = False
    for stats in sheet_stats:
        if stats["rows_with_extra_values"]:
            any_issue = True
            sample = ", ".join(str(row) for row in stats["rows_with_extra_values"][:20])
            lines.append(f"- `{stats['sheet']}` 存在未命名列内容，行：{sample}")
        if stats["blank_name_rows"]:
            any_issue = True
            sample = ", ".join(str(row) for row in stats["blank_name_rows"][:20])
            lines.append(f"- `{stats['sheet']}` 存在空名称行，行：{sample}")
    if not any_issue:
        lines.append("- 未发现")

    lines.extend(
        [
            "",
            "## 输出文件",
            "",
            "- `data/cards_raw/all_cards.jsonl`",
            "- `data/cards_raw/combat_characters.jsonl`",
            "- `data/cards_raw/attached_characters.jsonl`",
            "- `data/cards_raw/items.jsonl`",
            "- `data/cards_raw/titles.jsonl`",
            "- `data/cards_raw/scenes.jsonl`",
            "- `data/cards_raw/deprecated.jsonl`",
            "- `data/cards_raw/manifest.json`",
            "",
        ]
    )

    REPORT.write_text("\n".join(lines), encoding="utf-8")
    print(str(REPORT))


if __name__ == "__main__":
    main()
