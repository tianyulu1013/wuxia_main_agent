from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "outputs" / "2025-excel-sync-candidate" / "已制作_2025日志同步候选_PSD校准.xlsx"
OUT_DIR = ROOT / "data" / "cards_current"
DB_PATH = ROOT / "data" / "cards.sqlite"
REPORT = ROOT / "docs" / "current-card-database-report.md"
AUTHOR_OVERRIDES = ROOT / "data" / "author_ability_overrides.json"
FIELD_OVERRIDES = ROOT / "data" / "card_field_overrides.json"


SHEET_KEYS = {
    "战斗人物": "combat_characters",
    "附加人物": "attached_characters",
    "物品": "items",
    "称号": "titles",
    "场景": "scenes",
    "废弃": "deprecated",
}


FIELD_KEYS = {
    "名称": "title",
    "血量": "life",
    "身份": "identity",
    "描述": "description",
    "关系": "relationships",
    "兵器": "weapons",
    "出处": "source_work",
    "作者": "author_group",
    "性别": "gender",
    "特性": "traits",
    "类别": "item_category",
}


CARD_COLUMNS = [
    "title",
    "life",
    "identity",
    "description",
    "relationships",
    "weapons",
    "source_work",
    "author_group",
    "gender",
    "traits",
    "item_category",
]


ABILITY_KINDS = {"内功", "招式", "武功", "技能", "符卡", "*", "字", "说明"}
TYPED_ABILITY_KINDS = {"内功", "招式", "武功", "技能", "符卡"}
ABILITY_AUDIT = ROOT / "docs" / "ability-structure-audit.md"
NESTED_PARENT_HINTS = (
    "以下姿态",
    "轮换使用",
    "四大绝招",
    "二者共同使用",
    "全具有",
    "与敌战斗",
    "不中即死",
)
NESTED_PARENT_NAMES = {"三道真气"}
KNOWN_UNNAMED_ABILITY_CARDS = {"五个人头"}


def load_author_overrides() -> dict[str, Any]:
    if not AUTHOR_OVERRIDES.exists():
        return {}
    return json.loads(AUTHOR_OVERRIDES.read_text(encoding="utf-8"))


def load_field_overrides() -> dict[str, Any]:
    if not FIELD_OVERRIDES.exists():
        return {}
    return json.loads(FIELD_OVERRIDES.read_text(encoding="utf-8"))


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n").strip()
    return value


def is_blank(value: Any) -> bool:
    value = clean_cell(value)
    return value is None or value == ""


def normalize_title(title: str | None) -> str:
    if not title:
        return ""
    normalized = title.replace("（", "(").replace("）", ")")
    normalized = re.sub(r"\s+", "", normalized)
    return normalized


def stable_id(sheet: str, row_number: int, title: str | None) -> str:
    seed = f"{sheet}|{row_number}|{title or ''}"
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:12]
    return f"card_{digest}"


def stringify(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


def card_all_text(record: dict[str, Any]) -> str:
    parts: list[str] = []
    for key in CARD_COLUMNS:
        value = record["fields"].get(key)
        if value is not None:
            parts.append(str(value))
    return "\n".join(parts)


def explicit_ability_kind(line: str) -> str:
    text = line.lstrip()
    match = re.match(r"^(?:\d+[.．、]\s*)?(内功|招式|武功|技能|符卡)：", text)
    if match:
        return match.group(1)
    if text.startswith("*"):
        return "*"
    return ""


def split_ability_name(line: str) -> tuple[str, str, str]:
    text = line or ""
    heading = re.match(r"^(\s*)(?:\d+[.．、]\s*)?((?:内功|招式|武功|技能|符卡)：)", text)
    if heading:
        type_prefix = f"{heading.group(1)}{heading.group(2)}"
        rest = text[heading.end() :]
        exclusive = re.match(r"^(\s*【[^】]+】[:：])", rest)
        if exclusive:
            return type_prefix, exclusive.group(1), rest[len(exclusive.group(1)) :]
        name = re.match(r"^(\s*[^：:\n]{1,24}[:：])", rest)
        if name:
            return type_prefix, name.group(1), rest[len(name.group(1)) :]
        dash_name = re.match(r"^(\s*[^：:\n]{1,24}?)(?=\s+[-—－]{1,}|--)", rest)
        if dash_name:
            return type_prefix, dash_name.group(1), rest[len(dash_name.group(1)) :]
        return type_prefix, "", rest

    exclusive = re.match(r"^(\s*【[^】]+】[:：]?)", text)
    if exclusive:
        return "", exclusive.group(1), text[len(exclusive.group(1)) :]

    star = re.match(r"^(\s*\*[^:：\n]{0,80}[:：])", text)
    if star:
        return "", star.group(1), text[len(star.group(1)) :]

    name = re.match(r"^(\s*[^：:\n]{1,24}[:：])", text)
    if name:
        return "", name.group(1), text[len(name.group(1)) :]

    return "", "", text


def normalized_ability_name(raw_name: str) -> str | None:
    name = raw_name.strip()
    if not name:
        return None
    name = name.removesuffix("：").removesuffix(":").strip()
    name = name.removeprefix("*").strip()
    return name or None


def line_has_ability_name(line: str) -> bool:
    _, name, _ = split_ability_name(line)
    return bool(name)


def raw_name_has_colon(raw_name: str | None) -> bool:
    return bool(raw_name and (raw_name.rstrip().endswith(":") or raw_name.rstrip().endswith("：")))


def ability_is_exclusive_name(name: object) -> bool:
    return bool(name and "【" in str(name) and "】" in str(name))


def ability_text_is_identity(text: object) -> bool:
    return bool(re.search(r"[（(]身份[）)]\s*$", str(text or "").strip()))


def infer_line_kind(line: str, inherited_kind: str) -> tuple[str, list[str]]:
    flags: list[str] = []
    explicit = explicit_ability_kind(line)
    if explicit:
        return explicit, flags
    if line_has_ability_name(line):
        if inherited_kind in TYPED_ABILITY_KINDS:
            flags.append("inherited_kind")
            return inherited_kind, flags
        flags.append("implicit_word")
        return "字", flags
    if inherited_kind:
        flags.append("continuation_line")
        return inherited_kind, flags
    flags.append("free_text")
    return "说明", flags


def line_indent_width(line: str) -> int:
    return len(line) - len(line.lstrip())


def ability_allows_nested_named_lines(ability: dict[str, Any]) -> bool:
    if "typed_line_without_name" in ability.get("review_flags", []):
        return True
    if ability.get("name") in NESTED_PARENT_NAMES:
        return True
    text = "\n".join(ability.get("lines", []))
    return any(hint in text for hint in NESTED_PARENT_HINTS)


def override_matches(record: dict[str, Any], ability: dict[str, Any], match: dict[str, Any]) -> bool:
    source = record["source"]
    if "source_sheet" in match and source["sheet"] != match["source_sheet"]:
        return False
    if "source_row" in match and int(source["row"]) != int(match["source_row"]):
        return False
    if "title" in match and record.get("title") != match["title"]:
        return False
    if "ordinal" in match and int(ability["ordinal"]) != int(match["ordinal"]):
        return False
    return True


def record_override_matches(record: dict[str, Any], match: dict[str, Any]) -> bool:
    source = record["source"]
    if "source_sheet" in match and source["sheet"] != match["source_sheet"]:
        return False
    if "source_row" in match and int(source["row"]) != int(match["source_row"]):
        return False
    if "title" in match and record.get("title") != match["title"]:
        return False
    return True


def apply_field_overrides(record: dict[str, Any], overrides: dict[str, Any]) -> None:
    if not overrides:
        return
    for update in overrides.get("field_updates", []):
        if not record_override_matches(record, update.get("match", {})):
            continue
        for key, value in update.get("set", {}).items():
            record["fields"][key] = value
        for key, value in update.get("raw_set", {}).items():
            record["raw_fields"][key] = value


def normalize_ability_ids(record: dict[str, Any], abilities: list[dict[str, Any]]) -> None:
    for ordinal, ability in enumerate(abilities, start=1):
        ability["ordinal"] = ordinal
        ability["id"] = f"{record['id']}::ability::{ordinal:03d}"
        ability["is_exclusive"] = ability_is_exclusive_name(ability.get("name"))
        ability["is_identity"] = ability_text_is_identity(ability.get("text"))


def apply_author_overrides(
    record: dict[str, Any],
    abilities: list[dict[str, Any]],
    overrides: dict[str, Any],
) -> list[dict[str, Any]]:
    if not overrides:
        return abilities

    for update in overrides.get("ability_updates", []):
        match = update.get("match", {})
        for ability in abilities:
            if not override_matches(record, ability, match):
                continue
            for key, value in update.get("set", {}).items():
                ability[key] = value
            flags = set(ability.get("review_flags", []))
            flags.difference_update(update.get("remove_flags", []))
            flags.update(update.get("add_flags", []))
            ability["review_flags"] = sorted(flags)
            if "name" in update.get("set", {}):
                ability["is_exclusive"] = ability_is_exclusive_name(ability.get("name"))

    deleted: set[tuple[str, int]] = set()
    for delete in overrides.get("ability_deletes", []):
        match = delete.get("match", {})
        for ability in abilities:
            if override_matches(record, ability, match):
                deleted.add((ability["card_id"], int(ability["ordinal"])))

    split_matches = overrides.get("ability_splits", [])
    result: list[dict[str, Any]] = []
    for ability in abilities:
        ability_key = (ability["card_id"], int(ability["ordinal"]))
        if ability_key in deleted:
            continue

        split = next(
            (
                item
                for item in split_matches
                if override_matches(record, ability, item.get("match", {}))
            ),
            None,
        )
        if not split:
            result.append(ability)
            continue

        base_start = int(ability["start_line"])
        for offset, entry in enumerate(split.get("entries", [])):
            new_ability = {**ability}
            new_ability.update(
                {
                    "kind": entry.get("kind", ability["kind"]),
                    "name": entry.get("name"),
                    "raw_name": entry.get("raw_name"),
                    "type_prefix": entry.get("type_prefix"),
                    "text": entry.get("text", ""),
                    "start_line": base_start + offset,
                    "end_line": base_start + offset,
                    "is_exclusive": ability_is_exclusive_name(entry.get("name")),
                    "is_identity": ability_text_is_identity(entry.get("text")),
                    "owner_units": entry.get("owner_units"),
                    "owner_identity": entry.get("owner_identity"),
                    "owner_weapons": entry.get("owner_weapons"),
                    "review_flags": sorted(set(entry.get("flags", []))),
                }
            )
            result.append(new_ability)

    for confirmed in overrides.get("structure_confirmed", []):
        for ability in result:
            if not override_matches(record, ability, confirmed):
                continue
            flags = set(ability.get("review_flags", []))
            flags.add("author_confirmed_structure")
            ability["review_flags"] = sorted(flags)

    normalize_ability_ids(record, result)
    return result


def parse_abilities(record: dict[str, Any]) -> list[dict[str, Any]]:
    description = record["fields"].get("description")
    if not isinstance(description, str) or not description.strip():
        return []

    abilities: list[dict[str, Any]] = []
    inherited_kind = ""
    current: dict[str, Any] | None = None
    ordinal = 0
    physical_line = 0

    def finish_current() -> None:
        nonlocal current
        if not current:
            return
        current["text"] = "\n".join(current.pop("lines")).strip()
        current["is_identity"] = ability_text_is_identity(current["text"])
        current.pop("_nested_child_indent", None)
        abilities.append(current)
        current = None

    paragraphs = re.split(r"\n\s*\n", description.replace("\r\n", "\n").replace("\r", "\n").strip())
    for paragraph in paragraphs:
        if not paragraph.strip():
            inherited_kind = ""
            finish_current()
            continue

        inherited_kind = ""
        for line in paragraph.split("\n"):
            physical_line += 1
            if not line.strip():
                inherited_kind = ""
                finish_current()
                continue

            explicit = explicit_ability_kind(line)
            kind, flags = infer_line_kind(line, inherited_kind)
            type_prefix, raw_name, _ = split_ability_name(line)
            name = normalized_ability_name(raw_name)
            starts_named_ability = bool(explicit or name)
            line_is_indented = line != line.lstrip()
            line_indent = line_indent_width(line)
            append_typed_nested_ability = False
            if (
                current is not None
                and name
                and not explicit
                and kind == current["kind"]
                and current["kind"] in TYPED_ABILITY_KINDS
                and ability_allows_nested_named_lines(current)
            ):
                child_indent = current.get("_nested_child_indent")
                if child_indent is None:
                    current["_nested_child_indent"] = line_indent
                    append_typed_nested_ability = True
                elif child_indent == line_indent:
                    append_typed_nested_ability = True

            append_to_nested_ability = (
                current is not None
                and not explicit
                and (
                    append_typed_nested_ability
                    or
                    (
                        current["kind"] == "*"
                        and ability_allows_nested_named_lines(current)
                    )
                    or (
                        line_is_indented
                        and name
                        and current["kind"] == kind
                        and current["kind"] not in TYPED_ABILITY_KINDS
                    )
                    or (
                        not name
                        and current["kind"] not in TYPED_ABILITY_KINDS
                    )
                )
            )

            if append_to_nested_ability:
                nested_flags = [*flags]
                if name:
                    nested_flags.append("nested_named_line")
                    if not line_is_indented:
                        nested_flags.append("nested_named_line_without_indent")
                else:
                    nested_flags.append("nested_continuation_line")
                if line_is_indented:
                    nested_flags.append("nested_indented_line")
                current["lines"].append(line)
                current["end_line"] = physical_line
                current["review_flags"] = sorted(set([*current["review_flags"], *nested_flags]))
                continue

            if starts_named_ability or current is None or current["kind"] != kind:
                finish_current()
                ordinal += 1
                review_flags = list(flags)
                if explicit and not name:
                    if record.get("title") in KNOWN_UNNAMED_ABILITY_CARDS:
                        review_flags.append("known_unnamed_ability")
                    elif explicit != "*":
                        review_flags.append("typed_line_without_name")
                    else:
                        review_flags.append("unnamed_star_line")
                if explicit and explicit != "*" and name and not raw_name_has_colon(raw_name):
                    review_flags.append("typed_name_without_colon")
                if kind == "字" and raw_name.strip().startswith("【"):
                    review_flags.append("exclusive_word")
                if kind in TYPED_ABILITY_KINDS and not explicit and name:
                    review_flags.append("inherited_named_ability")
                    if line == line.lstrip():
                        review_flags.append("missing_indent_for_inherited")
                if kind == "说明" and name:
                    review_flags.append("named_free_text")
                if kind == "字" and not explicit and name and line != line.lstrip():
                    review_flags.append("indented_implicit_word")
                current = {
                    "id": f"{record['id']}::ability::{ordinal:03d}",
                    "card_id": record["id"],
                    "ordinal": ordinal,
                    "kind": kind,
                    "name": name,
                    "raw_name": raw_name.strip() or None,
                    "type_prefix": type_prefix.strip() or None,
                    "source_field": "description",
                    "start_line": physical_line,
                    "end_line": physical_line,
                    "lines": [line],
                    "is_exclusive": ability_is_exclusive_name(name),
                    "is_identity": ability_text_is_identity(line),
                    "owner_units": None,
                    "owner_identity": None,
                    "owner_weapons": None,
                    "review_flags": review_flags,
                }
            else:
                current["lines"].append(line)
                current["end_line"] = physical_line
                current["review_flags"] = sorted(set([*current["review_flags"], *flags]))

            if explicit in TYPED_ABILITY_KINDS:
                inherited_kind = explicit

        finish_current()
        inherited_kind = ""

    return abilities


def read_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, source: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    author_overrides = load_author_overrides()
    field_overrides = load_field_overrides()
    header_values = [clean_cell(cell.value) for cell in ws[1]]
    field_names: list[str] = []
    unnamed_count = 0
    for idx, header in enumerate(header_values, start=1):
        if isinstance(header, str) and header:
            field_names.append(header)
        else:
            unnamed_count += 1
            field_names.append(f"__unnamed_{idx}")

    records: list[dict[str, Any]] = []
    rows_with_extra_values: list[int] = []
    blank_name_rows: list[int] = []

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [clean_cell(value) for value in row]
        if all(is_blank(value) for value in values):
            continue

        raw_fields = {
            field_names[i]: values[i] if i < len(values) else None
            for i in range(len(field_names))
        }
        raw_fields = {key: value for key, value in raw_fields.items() if not is_blank(value)}

        fields: dict[str, Any] = {}
        extra_fields: dict[str, Any] = {}
        for key, value in raw_fields.items():
            mapped = FIELD_KEYS.get(key)
            if mapped:
                fields[mapped] = value
            else:
                extra_fields[key] = value

        title_value = fields.get("title")
        title = title_value if isinstance(title_value, str) else str(title_value) if title_value is not None else None
        if not title:
            blank_name_rows.append(row_number)
            continue

        unnamed_values = {
            key: value
            for key, value in raw_fields.items()
            if key.startswith("__unnamed_") and not is_blank(value)
        }
        if unnamed_values:
            rows_with_extra_values.append(row_number)

        record = {
            "id": stable_id(ws.title, row_number, title),
            "source": {
                "workbook": source.name,
                "sheet": ws.title,
                "row": row_number,
            },
            "category": SHEET_KEYS.get(ws.title, ws.title),
            "title": title,
            "normalized_title": normalize_title(title),
            "fields": fields,
            "extra_fields": extra_fields,
            "raw_fields": raw_fields,
        }
        apply_field_overrides(record, field_overrides)
        record["all_text"] = card_all_text(record)
        record["abilities"] = apply_author_overrides(record, parse_abilities(record), author_overrides)
        records.append(record)

    stats = {
        "sheet": ws.title,
        "category": SHEET_KEYS.get(ws.title, ws.title),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header_values": header_values,
        "record_count": len(records),
        "unnamed_header_count": unnamed_count,
        "rows_with_extra_values": rows_with_extra_values,
        "blank_name_rows": blank_name_rows,
    }
    return records, stats


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as fh:
        for record in records:
            fh.write(json.dumps(record, ensure_ascii=False, default=str))
            fh.write("\n")


def build_sqlite(path: Path, records: list[dict[str, Any]], source: Path, imported_at: str) -> None:
    if path.exists():
        path.unlink()

    conn = sqlite3.connect(path)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute(
            """
            CREATE TABLE metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE cards (
                id TEXT PRIMARY KEY,
                category TEXT NOT NULL,
                title TEXT,
                normalized_title TEXT,
                life TEXT,
                identity TEXT,
                description TEXT,
                relationships TEXT,
                weapons TEXT,
                source_work TEXT,
                author_group TEXT,
                gender TEXT,
                traits TEXT,
                item_category TEXT,
                source_workbook TEXT NOT NULL,
                source_sheet TEXT NOT NULL,
                source_row INTEGER NOT NULL,
                all_text TEXT NOT NULL,
                fields_json TEXT NOT NULL,
                raw_fields_json TEXT NOT NULL,
                extra_fields_json TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE card_abilities (
                id TEXT PRIMARY KEY,
                card_id TEXT NOT NULL REFERENCES cards(id) ON DELETE CASCADE,
                ordinal INTEGER NOT NULL,
                kind TEXT NOT NULL,
                name TEXT,
                raw_name TEXT,
                type_prefix TEXT,
                source_field TEXT NOT NULL,
                start_line INTEGER NOT NULL,
                end_line INTEGER NOT NULL,
                text TEXT NOT NULL,
                is_exclusive INTEGER NOT NULL,
                is_identity INTEGER NOT NULL,
                owner_units_json TEXT,
                owner_identity TEXT,
                owner_weapons_json TEXT,
                review_flags_json TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE VIRTUAL TABLE cards_fts USING fts5(
                card_id UNINDEXED,
                title,
                category UNINDEXED,
                description,
                relationships,
                all_text,
                tokenize='unicode61'
            )
            """
        )
        conn.executemany(
            "INSERT INTO metadata(key, value) VALUES (?, ?)",
            [
                ("source_workbook", source.name),
                ("source_path", str(source.relative_to(ROOT))),
                ("imported_at_utc", imported_at),
                ("record_count", str(len(records))),
            ],
        )

        card_rows = []
        ability_rows = []
        fts_rows = []
        for record in records:
            fields = record["fields"]
            source_info = record["source"]
            card_rows.append(
                (
                    record["id"],
                    record["category"],
                    record.get("title"),
                    record.get("normalized_title"),
                    *[stringify(fields.get(column)) for column in CARD_COLUMNS if column != "title"],
                    source_info["workbook"],
                    source_info["sheet"],
                    source_info["row"],
                    record["all_text"],
                    json.dumps(fields, ensure_ascii=False, default=str),
                    json.dumps(record["raw_fields"], ensure_ascii=False, default=str),
                    json.dumps(record["extra_fields"], ensure_ascii=False, default=str),
                )
            )
            fts_rows.append(
                (
                    record["id"],
                    record.get("title") or "",
                    record["category"],
                    stringify(fields.get("description")) or "",
                    stringify(fields.get("relationships")) or "",
                    record["all_text"],
                )
            )
            for ability in record.get("abilities", []):
                ability_rows.append(
                    (
                        ability["id"],
                        ability["card_id"],
                        ability["ordinal"],
                        ability["kind"],
                        ability.get("name"),
                        ability.get("raw_name"),
                        ability.get("type_prefix"),
                        ability["source_field"],
                        ability["start_line"],
                        ability["end_line"],
                        ability["text"],
                        1 if ability.get("is_exclusive") else 0,
                        1 if ability.get("is_identity") else 0,
                        json.dumps(ability.get("owner_units"), ensure_ascii=False) if ability.get("owner_units") is not None else None,
                        ability.get("owner_identity"),
                        json.dumps(ability.get("owner_weapons"), ensure_ascii=False) if ability.get("owner_weapons") is not None else None,
                        json.dumps(ability.get("review_flags", []), ensure_ascii=False),
                    )
                )

        conn.executemany(
            """
            INSERT INTO cards (
                id, category, title, normalized_title, life, identity, description,
                relationships, weapons, source_work, author_group, gender, traits, item_category,
                source_workbook, source_sheet, source_row, all_text,
                fields_json, raw_fields_json, extra_fields_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            card_rows,
        )
        conn.executemany(
            """
            INSERT INTO cards_fts(card_id, title, category, description, relationships, all_text)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            fts_rows,
        )
        conn.executemany(
            """
            INSERT INTO card_abilities (
                id, card_id, ordinal, kind, name, raw_name, type_prefix, source_field,
                start_line, end_line, text, is_exclusive, is_identity,
                owner_units_json, owner_identity, owner_weapons_json, review_flags_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ability_rows,
        )
        conn.execute("CREATE INDEX idx_cards_title ON cards(title)")
        conn.execute("CREATE INDEX idx_cards_normalized_title ON cards(normalized_title)")
        conn.execute("CREATE INDEX idx_cards_category ON cards(category)")
        conn.execute("CREATE INDEX idx_cards_author_group ON cards(author_group)")
        conn.execute("CREATE INDEX idx_cards_source_sheet_row ON cards(source_sheet, source_row)")
        conn.execute("CREATE INDEX idx_abilities_card_id ON card_abilities(card_id)")
        conn.execute("CREATE INDEX idx_abilities_kind ON card_abilities(kind)")
        conn.execute("CREATE INDEX idx_abilities_name ON card_abilities(name)")
        conn.commit()
    finally:
        conn.close()


def write_report(
    source: Path,
    records: list[dict[str, Any]],
    sheet_stats: list[dict[str, Any]],
    imported_at: str,
) -> None:
    by_category = Counter(record["category"] for record in records)
    by_author = Counter(record["fields"].get("author_group") for record in records if record["fields"].get("author_group"))
    title_counter = Counter(record["title"] for record in records if isinstance(record.get("title"), str))
    duplicates = {title: count for title, count in title_counter.items() if count > 1}

    duplicate_locations: dict[str, list[str]] = defaultdict(list)
    for record in records:
        title = record.get("title")
        if title in duplicates:
            source_info = record["source"]
            duplicate_locations[title].append(f"{source_info['sheet']}!{source_info['row']}")

    lines = [
        "# 当前卡牌数据库导入报告",
        "",
        f"- 来源 Excel：`{source.relative_to(ROOT)}`",
        f"- SQLite：`{DB_PATH.relative_to(ROOT)}`",
        f"- JSONL：`{OUT_DIR.relative_to(ROOT)}/all_cards.jsonl`",
        f"- 导入时间 UTC：`{imported_at}`",
        f"- 总记录数：{len(records)}",
        "",
        "## Sheet 统计",
        "",
        "| Sheet | 分类 | 记录数 | 最大行 | 最大列 | 未命名表头数 | 有未命名列内容行数 | 空名称行数 |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for stats in sheet_stats:
        lines.append(
            f"| {stats['sheet']} | `{stats['category']}` | {stats['record_count']} | "
            f"{stats['max_row']} | {stats['max_column']} | {stats['unnamed_header_count']} | "
            f"{len(stats['rows_with_extra_values'])} | {len(stats['blank_name_rows'])} |"
        )

    lines.extend(["", "## 分类统计", ""])
    for category, count in sorted(by_category.items()):
        lines.append(f"- `{category}`: {count}")

    lines.extend(["", "## 作者统计", ""])
    for author, count in by_author.most_common():
        lines.append(f"- {author}: {count}")

    lines.extend(["", "## 重名标题", ""])
    if duplicate_locations:
        for title, locations in sorted(duplicate_locations.items()):
            lines.append(f"- {title}: {', '.join(locations)}")
    else:
        lines.append("- 无")

    lines.extend(
        [
            "",
            "## 查询示例",
            "",
            "```sql",
            "SELECT title, category, source_sheet, source_row FROM cards WHERE normalized_title = '周芷若';",
            "```",
            "",
            "```sql",
            "SELECT title, category, source_sheet, source_row FROM cards WHERE all_text LIKE '%不利%' LIMIT 20;",
            "```",
            "",
        ]
    )

    REPORT.write_text("\n".join(lines), encoding="utf-8")


def flatten_abilities(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        source = record["source"]
        for ability in record.get("abilities", []):
            rows.append(
                {
                    **ability,
                    "card_title": record.get("title"),
                    "card_category": record.get("category"),
                    "source_sheet": source["sheet"],
                    "source_row": source["row"],
                    "source_work": record["fields"].get("source_work"),
                    "author_group": record["fields"].get("author_group"),
                }
            )
    return rows


def write_ability_audit(records: list[dict[str, Any]], abilities: list[dict[str, Any]], imported_at: str) -> None:
    by_kind = Counter(ability["kind"] for ability in abilities)
    by_flag = Counter(flag for ability in abilities for flag in ability.get("review_flags", []))
    review_items = [
        ability
        for ability in abilities
        if any(
            flag in ability.get("review_flags", [])
            for flag in [
                "inherited_named_ability",
                "implicit_word",
                "typed_line_without_name",
                "typed_name_without_colon",
                "nested_named_line_without_indent",
                "missing_indent_for_inherited",
                "indented_implicit_word",
                "exclusive_word",
                "free_text",
            ]
        )
    ]

    def location(ability: dict[str, Any]) -> str:
        return f"{ability['source_sheet']}!{ability['source_row']}:{ability['start_line']}"

    def one_line(text: str, limit: int = 180) -> str:
        compacted = re.sub(r"\s+", " ", text).strip()
        return compacted if len(compacted) <= limit else f"{compacted[:limit]}..."

    def section_for(flag: str, title: str, limit: int = 80) -> list[str]:
        items = [ability for ability in review_items if flag in ability.get("review_flags", [])]
        lines = ["", f"## {title}", "", f"- 数量：{len(items)}"]
        for ability in items[:limit]:
            lines.extend(
                [
                    "",
                    f"### {ability.get('card_title')} / {location(ability)}",
                    "",
                    f"- 当前判断：`{ability['kind']}` / 名称：`{ability.get('name') or '—'}`",
                    f"- 标记：`{', '.join(ability.get('review_flags', []))}`",
                    f"- 原文：{one_line(ability['text'])}",
                ]
            )
        if len(items) > limit:
            lines.append(f"\n- 仅显示前 {limit} 条，完整列表见 `data/cards_current/abilities.jsonl`。")
        return lines

    lines = [
        "# 特技结构审计报告",
        "",
        "本报告来自当前数据库的 `description` 字段自动解析。它不会替代作者裁定，只用于把疑似结构问题集中列出来。",
        "",
        f"- 导入时间 UTC：`{imported_at}`",
        f"- 卡牌数：{len(records)}",
        f"- 抽取特技/说明块数：{len(abilities)}",
        f"- 结构化特技表：`data/cards.sqlite` / `card_abilities`",
        f"- JSONL：`data/cards_current/abilities.jsonl`",
        "",
        "## 类型统计",
        "",
    ]
    for kind, count in by_kind.most_common():
        lines.append(f"- `{kind}`: {count}")

    lines.extend(["", "## 审计标记统计", ""])
    if by_flag:
        for flag, count in by_flag.most_common():
            lines.append(f"- `{flag}`: {count}")
    else:
        lines.append("- 无")

    lines.extend(
        section_for(
            "inherited_named_ability",
            "无前缀但继承上一类型的特技",
            100,
        )
    )
    lines.extend(section_for("missing_indent_for_inherited", "继承上一类型但没有缩进", 120))
    lines.extend(section_for("indented_implicit_word", "有缩进但按字处理", 120))
    lines.extend(section_for("typed_name_without_colon", "有类型前缀但特技名缺少冒号", 120))
    lines.extend(section_for("nested_named_line_without_indent", "特技子项没有缩进", 120))
    lines.extend(section_for("implicit_word", "无前缀且按字处理的特技", 80))
    lines.extend(section_for("exclusive_word", "带【】的字", 80))
    lines.extend(section_for("typed_line_without_name", "有类型前缀但未识别出特技名", 80))
    lines.extend(section_for("free_text", "说明性文本块", 80))

    ABILITY_AUDIT.write_text("\n".join(lines), encoding="utf-8")


def build(source: Path) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    imported_at = datetime.now(timezone.utc).isoformat()
    wb = openpyxl.load_workbook(source, data_only=False)

    all_records: list[dict[str, Any]] = []
    sheet_stats: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        records, stats = read_sheet(ws, source)
        all_records.extend(records)
        sheet_stats.append(stats)
        key = SHEET_KEYS.get(ws.title, ws.title)
        write_jsonl(OUT_DIR / f"{key}.jsonl", records)

    write_jsonl(OUT_DIR / "all_cards.jsonl", all_records)
    all_abilities = flatten_abilities(all_records)
    write_jsonl(OUT_DIR / "abilities.jsonl", all_abilities)
    (OUT_DIR / "manifest.json").write_text(
        json.dumps(
            {
                "source_workbook": source.name,
                "source_path": str(source.relative_to(ROOT)),
                "imported_at_utc": imported_at,
                "record_count": len(all_records),
                "ability_count": len(all_abilities),
                "sheets": sheet_stats,
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )
    build_sqlite(DB_PATH, all_records, source, imported_at)
    write_report(source, all_records, sheet_stats, imported_at)
    write_ability_audit(all_records, all_abilities, imported_at)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the queryable card database from the current Excel workbook.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    args = parser.parse_args()

    source = args.source
    if not source.is_absolute():
        source = ROOT / source
    if not source.exists():
        raise FileNotFoundError(source)
    build(source)
    print(str(DB_PATH))
    print(str(OUT_DIR / "all_cards.jsonl"))
    print(str(OUT_DIR / "abilities.jsonl"))
    print(str(REPORT))
    print(str(ABILITY_AUDIT))


if __name__ == "__main__":
    main()
