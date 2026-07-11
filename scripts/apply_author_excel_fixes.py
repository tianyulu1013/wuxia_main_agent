from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "2025-excel-sync-candidate" / "已制作_2025日志同步候选_PSD校准.xlsx"
BACKUP_DIR = ROOT / "outputs" / "2025-excel-sync-candidate" / "backups"


FIXES = [
    {
        "sheet": "战斗人物",
        "title": "浪翻云",
        "field": "描述",
        "old": "*唯能极于情，故能极于剑，覆雨剑法进入",
        "new": "*唯能极于情，故能极于剑：覆雨剑法进入",
    },
    {
        "sheet": "战斗人物",
        "title": "风清扬",
        "field": "描述",
        "old": "招式：独孤九剑 -- 1000",
        "new": "招式：独孤九剑：-- 1000",
    },
    {
        "sheet": "战斗人物",
        "title": "风清扬",
        "field": "描述",
        "old": "技能：华山剑宗 学会可见剑法",
        "new": "技能：华山剑宗：学会可见剑法",
    },
    {
        "sheet": "战斗人物",
        "title": "无尘道长",
        "field": "描述",
        "old": "招式：七十二路夺命追魂剑  -- 300",
        "new": "招式：七十二路夺命追魂剑：-- 300",
    },
    {
        "sheet": "物品",
        "title": "辟邪剑谱",
        "field": "描述",
        "old": "招式：辟邪剑法 -- 900破",
        "new": "招式：辟邪剑法：-- 900破",
    },
    {
        "sheet": "战斗人物",
        "title": "安隆",
        "field": "描述",
        "old": "招式：天心莲环 击中敌",
        "new": "招式：天心莲环：击中敌",
    },
    {
        "sheet": "战斗人物",
        "title": "浦饭幽助",
        "field": "描述",
        "old": "招式：百裂拳 若对阵敌",
        "new": "招式：百裂拳：若对阵敌",
    },
    {
        "sheet": "战斗人物",
        "title": "李大娘",
        "field": "描述",
        "old": "招式：穿云掌 土- 500",
        "new": "招式：穿云掌：土- 500",
    },
    {
        "sheet": "战斗人物",
        "title": "李大娘",
        "field": "描述",
        "old": "技能：罗刹鬼婆 李大娘",
        "new": "技能：罗刹鬼婆：李大娘",
    },
]


def main() -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / f"{WORKBOOK.stem}.{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = openpyxl.load_workbook(WORKBOOK)
    applied: list[str] = []
    for fix in FIXES:
        ws = wb[fix["sheet"]]
        headers = {cell.value: cell.column for cell in ws[1]}
        title_col = headers["名称"]
        field_col = headers[fix["field"]]
        target_row = None
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=title_col).value
            if str(value or "").strip() == fix["title"]:
                target_row = row
                break
        if target_row is None:
            raise RuntimeError(f"Card not found: {fix['title']}")

        cell = ws.cell(row=target_row, column=field_col)
        value = str(cell.value or "")
        if fix["new"] in value:
            applied.append(f"{fix['title']}: already fixed at row {target_row}")
            continue
        if fix["old"] not in value:
            raise RuntimeError(f"Expected text not found for {fix['title']} row {target_row}")
        cell.value = value.replace(fix["old"], fix["new"], 1)
        applied.append(f"{fix['title']}: fixed {fix['field']} at row {target_row}")

    wb.save(WORKBOOK)
    print(f"backup={backup}")
    for item in applied:
        print(item)


if __name__ == "__main__":
    main()
